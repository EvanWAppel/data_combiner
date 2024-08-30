#import plotly.express as px
import pandas as pd
import os
import os.path
import re as re
import win32com.client
import datetime
import numpy as np
import time
import json 
from datetime import datetime, date, timedelta
import openpyxl as xl
#import parquet
import requests
from facebook_business.adobjects.adaccount import AdAccount
from facebook_business.api import FacebookAdsApi
from facebook_business.adobjects.campaign import Campaign 
from facebook_business.bootstrap import Authentication
from facebook_business.adobjects.ad import Ad
from facebook_business.adobjects.adsinsights import AdsInsights
from facebook_business.adobjects.objectparser import ObjectParser


class assemblers:

    def __init__(self):
        pass

    def data_assembler(self, source, criteria) -> list:
        targets = []
        for i in os.listdir(source):
            x = re.search(criteria,i)
            if x is not None:
                targets.append(i)
        return targets

    def data_maker(self,data,data_assembler: list,sheet: list="default",skiprows:str=None):
        
        dataframe = pd.DataFrame()
        if sheet != "default":
            for i in data_assembler:
                sheets = xl.load_workbook(data+"\\"+i).sheetnames
                if re.search(".xlsx|.xls",i) is not None:
                    sheet_table = pd.DataFrame()
                    try:
                        for j in sheets:
                            x = pd.read_excel(data+"\\"+i,sheet_name=j ,skiprows=skiprows)
                            sheet_table = ([sheet_table,x])
                        dataframe = pd.concat([dataframe,sheet_table])
                    except Exception as e:
                        print(e)
                elif re.search(".csv",i) is not None:
                    sheet_table = pd.DataFrame()
                    try:
                        for j in sheets:
                            x = pd.read_csv(data+"\\"+i,sheet_name=j,skiprows=skiprows)
                            sheet_table = ([sheet_table,x])
                        dataframe = pd.concat([dataframe,sheet_table])

                    except Exception as e:
                        print(e)
                else:
                    print(f"{time.time()} >>> Exotic data type, please add to data_maker.",file=open(log,'a'))
        else:
            for i in data_assembler:
                sheets = xl.load_workbook(data+"\\"+i).sheetnames
                if re.search(".xlsx|.xls",i) is not None:
                    sheet_table = pd.DataFrame()
                    try:
                        for j in sheets:

                            x = pd.read_excel(data+"\\"+i,skiprows=skiprows)
                            sheet_table = ([sheet_table,x])
                        dataframe = pd.concat([dataframe,sheet_table])
                    except Exception as e:
                        print(e)
                elif re.search(".csv",i) is not None:
                    sheet_table = pd.DataFrame()
                    try:
                        for j in sheets:
                            x = pd.read_csv(data+"\\"+i,skiprows=skiprows)
                            sheet_table = ([sheet_table,x])
                        dataframe = pd.concat([dataframe,sheet_table])
                    except Exception as e:
                        print(e)
                else:
                    print(f"{time.time()} >>> Exotic data type, please add to data_maker.",file=open(log,'a'))
        return(dataframe)
    
    def spon(self,source: str):
        cocod = data.c1b().coco(source=source)
        nascard = data.c1b().nascar(source=source)
        raidersd = data.c1b().raiders(source=source)
        vgkd = data.c1b().vgk(source=source)
        wwed = data.c1b().wwe(source=source)
        wsod = data.c1b().wso(source=source)
        collected = pd.concat([cocod,nascard,raidersd,vgkd,wwed,wsod])
        collected['DATE'] = pd.to_datetime(collected['DATE'])
        spend_file = pd.read_excel(source+r"\\Spend Record 2.xlsx")
        joined = pd.merge(collected,spend_file,on=['Format','Product'],how='left')
        filtered = joined[((joined['DATE']>=joined['START'])&(joined['DATE']<=joined['END']))|(joined['START'].isna())]
        grouped = filtered.groupby(['Format','Product','START','END','AMT SPEND'],as_index=False).agg({'IMPRESSIONS':"sum"})
        spon = pd.merge(collected,grouped,on=["Format","Product"],how='left')
        spon = spon[(spon['DATE']>=spon['START'])&(spon['DATE']<=spon['END'])|(spon['START'].isna())]
        spon = spon.rename(columns={'IMPRESSIONS_x':'IMPRESSIONS',
                                    'IMPRESSIONS_y':'grouped_impressions',
                                    'AMT SPEND':'SPEND_TOTAL'})
        spon = spon.assign(spend = (spon['IMPRESSIONS'] / spon['grouped_impressions']) * spon['SPEND_TOTAL'])
        spon = spon.rename(columns={'IMPRESSIONS':'TOTAL_IMPRESSIONS',
                                                                    'SPEND_TOTAL':'grouped_spend',
                                                                    'spend':'TOTAL_SPEND'})
        spon = spon.rename(columns={'Product':'PRODUCT',
                                    'date':'DATE',
                                    'Format':'FORMAT',
                                    'Event':'EVENT',
                                    'TOTAL_IMPRESSIONS':'IMPRESSIONS',
                                    'TOTAL_SPEND':'SPEND'})
        spon = spon.assign(CLICKS=0)
        spon['CLICKS'] = spon['CLICKS'].astype('float64')
        spon['DATE'] = pd.to_datetime(spon['DATE'])
        spon = spon[[  'DATE'
                    ,'FORMAT'
                    ,'PRODUCT'
                    ,'EVENT'
                    ,'PLACEMENT'
                    ,'IMPRESSIONS'
                    ,'CLICKS'
                    ,'SPEND']]
        return spon
    
    def tableau(self):
        print(f"{time.time()} >>> Libraries loaded successfully...",file=open(log,'a'))
        start = time.perf_counter()
        ####DATA CLASS INITIALIZER###########################################################################
        try:
            d = data.assumptions()
            print(f"{time.time()} >>> Data class loaded...",file=open(log,'a'))
        except Exception as e:
            print(e)
        self.data_object = d
        end = time.perf_counter()
        print(str(end - start))
        # BENCHMARK 1.0399991879239678
        ####TELEVISION FUNCTION###############################################################################
        try:
            self.tv = data.dla().television()
            print(f"{time.time()} >>> TV data loaded...",file=open(log,'a'))
        except Exception as e:
            print(e)
        end = time.perf_counter()
        print(str(end - start))
        # BENCHMARK 32.12839179998264
        ####DIGITAL FUNCTION###################################################################################
        try:
            self.dig = data.dla().digital()
            print(f"{time.time()} >>> Digital data loaded...",file=open(log,'a'))
        except Exception as e:
            print(e)
        end = time.perf_counter()
        print(str(end - start))
        # BENCHMARK: 322.2016500000027
        ####SPONSORSHIP FUNCTION###############################################################################
        try:
            self.spond = assemblers().spon(source=data.assumptions().data)
            print(f"{time.time()} >>> Sponsorship data loaded...",file=open(log,'a'))
        except Exception as e:
            print(e)
        end = time.perf_counter()
        print(str(end - start))
        # BENCHMARK: 2.06392469999264
        ####RADIO FUNCTION####################################################################################
        try:
            self.rad = data.c1b().radio(d.data)
            print(f"{time.time()} >>> Radio data loaded...",file=open(log,'a'))
        except Exception as e:
            print(e)
        end = time.perf_counter()
        print(str(end - start))
        # BENCHMARK: 1.000
        ####PR FUNCTION######################################################################################
        try:
            self.prd = data.c1b().pr(d.data)
            print(f"{time.time()} >>> PR data loaded...",file=open(log,'a'))
        except Exception as e:
            print(e)
        end = time.perf_counter()
        print(str(end - start))
        # BENCHMARK: 2.5071819000004325
        ####SOCIAL FUNCTION####################################################################################
        try:
            self.soc = data().rudimentary_social()
            print(f"{time.time()} >>> Social data loaded...",file=open(log,'a'))
        except Exception as e:
            print(e)
        end = time.perf_counter()
        print(str(end - start))
        # BENCHMARK: 7.8032360000070184
        ####MLB Function####################################################################################
        try:
            self.mlb = data.c1b().mlb(d.data)
            print(f"{time.time()} >>> mlb data loaded...",file=open(log,'a'))
        except Exception as e:
            print(e)
        end = time.perf_counter()
        print(str(end - start))
        ####YOUTUBE ASSEMBLY################################################
        try:
            self.yt = data.dla().youtube()
            print(f"{time.time()} >>> Youtube Data Loaded...",file=open(log,'a'))
        except Exception as e:
            print(e)
        end = time.perf_counter()
        print(str(end - start))
        ####MASTER ASSEMBLY#################################################
        try:
            self.master = pd.concat([self.tv.reset_index(drop=True)
                                     ,self.prd.reset_index(drop=True)
                                     ,self.soc.reset_index(drop=True)
                                     ,self.spond.reset_index(drop=True)
                                     ,self.dig.reset_index(drop=True)
                                     ,self.rad.reset_index(drop=True)
                                     ,self.mlb.reset_index(drop=True)
                                     ,self.yt.reset_index(drop=True)])
            self.master_raw = self.master
            self.master = self.master[~self.master['DATE'].isna()]
            self.master = self.master[self.master['DATE']!=0]
            self.master['DATE'] = pd.to_datetime(self.master['DATE'])
            self.master = self.master.assign(Year=self.master['DATE'].dt.year)
            self.imp = pd.read_excel(d.data+r'\IMP FACTOR.xlsx')
            self.mastera = self.master[self.master['FORMAT']=='Sponsorship']
            self.mastera = pd.merge(self.mastera,self.imp,how='left',on=['FORMAT','PRODUCT','Year'])
            self.masterb = self.master[self.master['FORMAT']!='Sponsorship']
            self.impb = self.imp[['Year','Weight','Exposure','Error Rate','Qi Score','Notes','Error Rate Platform','imp factor','FORMAT']]
            self.masterb = pd.merge(self.masterb,self.impb,how='left',on=['FORMAT','Year'])
            self.master = pd.concat([self.mastera,self.masterb])
            self.master['PRODUCT'] = self.master['PRODUCT'].str.upper()
            print(f"{time.time()} >>> Master table assembled...",file=open(log,'a'))
        except Exception as e:
            print(e)
        ####Spend Assembly#####################################################################################################
        try:
            # Create list of all dates for each format
            formats = ['TV','Streaming','Social','Display','Audio','Radio','PR','Sponsorship']
            s_prods = [['Sponsorship','NASCAR'],['Sponsorship','COCO'],['Sponsorship','VGK'],['Sponsorship','WWE'],['Sponsorship','RAIDERS'],['Sponsorship','AVIATORS']]
            s_prods = pd.DataFrame(s_prods,columns = ['FORMAT','PRODUCT'])
            all_dates = pd.DataFrame()
        except Exception as e:
            print(e)
        try:
            for i in formats:
                x = pd.DataFrame(pd.date_range(start='1/1/2020', end='1/01/2025'))
                x = x.rename(columns ={ 0: 'DATE'})
                x = x.assign(FORMAT = i)
                all_dates = pd.concat([all_dates,x])
        except Exception as e:
            print(e)
        try:    
            all_dates = pd.merge(all_dates,s_prods,on='FORMAT',how='left')
        except Exception as e:
            print(e)
        try:
            all_dates = all_dates.fillna('blank')
        except Exception as e:
            print(e)
            # Create quarter column
        try:
            all_dates['quarter'] = 'Q'+pd.DatetimeIndex(all_dates['DATE']).quarter.astype('str')+' - '+pd.DatetimeIndex(all_dates['DATE']).year.astype('str')
        except Exception as e:
            print(e)
            # create quarter counter
        try:
            dc = all_dates.groupby(['quarter','FORMAT'],as_index=False).agg({'DATE':np.size})
        except Exception as e:
            print(e)
        try:
            all_dates = pd.merge(all_dates,dc,how='left',on=['quarter','FORMAT'])
        except Exception as e:
            print(e)
        try:
            all_dates = all_dates.rename(columns = {'DATE_x':'DATE',
                                                    'DATE_y':'counter'})
        except Exception as e:
            print(e)
            # Create table of known spend and impression data
        try:
            m = self.master.assign(PRODUCT = np.where(self.master['FORMAT']=='Sponsorship',self.master['PRODUCT'],(np.where(self.master['FORMAT']!='Sponsorship','blank','blank'))))
        except Exception as e:
            print(e)
        try:
            m = m.groupby(['FORMAT','PRODUCT','DATE'],as_index=False).agg({'SPEND':"sum",'IMPRESSIONS':"sum"})
        except Exception as e:
            print(e)
        try:
            m['quarter'] = 'Q'+pd.DatetimeIndex(m['DATE']).quarter.astype('str')+' - '+pd.DatetimeIndex(m['DATE']).year.astype('str')
        except Exception as e:
            print(e)
        try:
            m['DATE'] = pd.to_datetime(m['DATE'])
        except Exception as e:
            print(e)
            # Merge date table and known data table
        try:
            dist = pd.merge(all_dates,m,on=['DATE','FORMAT','PRODUCT'],how='left')
        except Exception as e:
            print(e)
        try:
            dist = dist.rename(columns = {'quarter_x':'quarter'})
        except Exception as e:
            print(e)
            # merge dist with budget
        try:
            budget = pd.read_excel(d.data+r'\BrandQuaterlyGoalsByChannel.xlsx')
        except Exception as e:
            print(e)
        try:
            table = pd.merge(dist,budget,on=['quarter','FORMAT'],how='left')
        except Exception as e:
            print(e+"HELLO?!",file=open(log,'a'))
            # Merge table with imp factor data
        try:
            self.imp = self.imp[['FORMAT','PRODUCT','Weight','Exposure','Error Rate Platform','Qi Score','imp factor','Year']]
        except Exception as e:
            print(e)
        try:
            table['Year'] = table['DATE'].dt.year
        except Exception as e:
            print(e)
        try:
            table = pd.merge(table,self.imp,how='left',on=['FORMAT','PRODUCT','Year'])
        except Exception as e:
            print(e)
        try:
            table = table.rename(columns={'FORMAT_x':'FORMAT',
                                        'PRODUCT_x':'PRODUCT'})
        except Exception as e:
            print(e)
        try:
            table = table.fillna(0)
        except Exception as e:
            print(e)
        try:
            # calculate daily budget allocation
            table = table.assign(daily_spend_target=table['Planned Spend']/table['counter'],
                                daily_net_imp_target=table['Planned Net Impressions']/table['counter'],
                                daily_gross_imp_target=table['Planned Gross Impressions']/table['counter'])
        except Exception as e:
            print(e)
        try:
            #Select only DATE, Format, SPEND, IMPRESSIONS, BUDGET
            table = table[['DATE'
                        ,'quarter'
                        ,'FORMAT'
                        ,'PRODUCT'
                        ,'SPEND'
                        ,'daily_spend_target'
                        ,'IMPRESSIONS'
                        ,'daily_net_imp_target'
                        ,'Planned Net CPM'
                        ,'daily_gross_imp_target'
                        ,'counter'
                        ,'imp factor']]
            # Filter out rows without data
        except Exception as e:
            print(e)
        try:
            table = table.assign(hasher=table['SPEND']+table['IMPRESSIONS']+table['daily_spend_target']+table['daily_net_imp_target'])
        except Exception as e:
            print(e)
        try:
            self.table = table[table['hasher']>0]
        except Exception as e:
            print(e)
        try:
            print(f"{time.time()} >>> Spend data table assembled...",file=open(log,'a'))
        except Exception as e:
            print(e)
        ####PRINTER#######################################################################################################################
        try:
            self.table.to_csv(d.product+r'\BrandBudget.csv')
            self.master.to_csv(d.product+r'\BrandSpendMaster.csv')
            print(f"{time.time()} >>> Brand Spend Data Model dashboard datasets have been updated.",file=open(log,'a'))
            print(f"{time.time()} >>> Thank you for your patience.",file=open(log,'a'))
        except Exception as e:
            print(e)
        end = time.perf_counter()
        print(str(end - start))
        ####ANCILLARY#######################################################################################################################
        ####Adelaide#######################################################################################################################
        try:
            data.dla().adelaide()
            print(f"{time.time()} >>> Adelaide Data pushed to Target Folder.",file=open(log,'a'))
        except:
            print(f"{time.time()} >>> Adelaide is not working...",file=open(log,'a'))
        end = time.perf_counter()
        print(str(end - start))
        ####VideoAmp#######################################################################################################################
        # try:
        #     data.dla().VideoAmpDataProcessor()
        #     print(f"{time.time()} >>> VideoAmp Data pushed to Target Folder.",file=open(log,'a'))
        # except:
        #     print(f"{time.time()} >>> VideoAmp is not working...",file=open(log,'a'))
        # end = time.perf_counter()
        # print(str(end - start))

    def file_writer(self,destination: str = None, tables:list = None, names:list = None):
        if len(tables) != len(names):
            print(f"{time.time()} >>> The length of files and tabs must be equal",file=open(log,'a'))
            exit()
        def write_excel(filename:str = None,sheetname: str =None,dataframe: pd.DataFrame = None):
            if os.path.isfile(filename) == True:
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a',if_sheet_exists='replace') as writer:
                    dataframe.to_excel(writer, sheet_name=sheetname,index=False)
            else: 
                with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
                    dataframe.to_excel(writer, sheet_name=sheetname,index=False)
        for i in tables:
            for j in names:
                write_excel(filename=destination,sheetname=j,dataframe=i)

        
class data:
    def __init__(self):
        pass
    class assumptions:
        
        def __init__(self):
            if os.path.isdir(r"N:\MARKET\Marketing Analytics\Brand Analysis\04 - Dashboards\Brand Spend Data Model"):
                self.root = r"N:\MARKET\Marketing Analytics\Brand Analysis\04 - Dashboards\Brand Spend Data Model"
            elif os.path.isdir(r"N:\Marketing Analytics\Brand Analysis\04 - Dashboards\Brand Spend Data Model"):
                self.root = r"N:\Marketing Analytics\Brand Analysis\04 - Dashboards\Brand Spend Data Model"
            else:
                self.root = input("Enter the file path for the Brand Spend Data Model")
                # modification: added MARKET to file path plus some other cool shit
            self.data = self.root+r"\Source Data"
            self.product = self.root+r"\Target Data"
            self.imp = pd.read_excel(self.data+r'\IMP FACTOR.xlsx')
            self.budget = pd.read_excel(self.data+r'\BrandQuaterlyGoalsByChannel.xlsx')

    class c1b:

        def __init__(self):
            pass

        def coco(self,source: str):
            coco = pd.read_excel(source+r"\COCO YTD.XLSX")
            co = coco[['DATE','Format','Events','Locations','IMPRESSIONS']]
            co = co.rename(columns={'Events':'Event',
                                    'Locations':'PLACEMENT'})
            co = co.assign(Product="COCO")
            return co
        
        def mlb(self,source: pd.DataFrame):
            mlba = pd.read_excel(source+r"\mlb.xlsx")
            spend2023 = 864690
            mlba = mlba.assign(IMPRESSIONS=mlba['Sponsorship Impressions (000)']*1000,
                            FORMAT = 'TV',
                            PRODUCT = 'MLB')
            mlba = mlba.assign(tot_imp = np.sum(mlba['IMPRESSIONS']))
            mlba = mlba.assign(SPEND = (spend2023/mlba['tot_imp'])*mlba['IMPRESSIONS'])
            return mlba
        
        def nascar(self,source: str):
            nascar = pd.read_excel(source+r"\Nascar YTD 22 OLD.XLSX",sheet_name='Nascar Race Level')
            na = nascar[['Events','Exposure Impressions (000)','DATE']]
            na = na.assign(Product="NASCAR")
            na = na.assign(Format='Sponsorship')
            na['DATE'] = pd.to_datetime(na['DATE'])
            na = na.assign(Event=na['Events'].str.extract(r'\d+\-\d+\s(.*)'))
            na = na.assign(IMPRESSIONS=na['Exposure Impressions (000)']*1000)
            return na
        
        def pr(self,source:str):
            # Find relevant files
            files = assemblers().data_assembler(source=source,criteria=r"PR Cision")
            # assemble files into dataframe
            df = assemblers().data_maker(source,files,sheet='Year')
            # Cast date into datetime format
            df['Date'] = pd.to_datetime(df['Date'])
            # Create new columns for month and year
            df = df.assign(month=df['Date'].dt.month,
                        year=df['Date'].dt.year)
            # Quality Mention Stuff
            # Get Quality mention data
            file = assemblers().data_assembler(source,r"QualityOutletMention")
            table = assemblers().data_maker(source,file,sheet=r'Quality Outlet Coverage')
            # List of links to look up
            regex_lst = table['Link'].astype('str').to_list()
            # Lowercasing
            regex_lst = [x.lower() for x in regex_lst]
            df['Link'] = df['Link'].fillna("null")
            df['quality_site'] = df['Link'].apply(lambda x: True if any(i in x for i in regex_lst) else False)
            # PR Impressions grouped
            pr_imps = df.groupby(['month','year'],as_index=False).agg({'SEO Impact':"sum"})
            # Spend Data
            file = assemblers().data_assembler(source,r"spend\.csv")
            pr_spend = assemblers().data_maker(source,file)
            pr_spend['Amount'] = pr_spend['Amount'].astype('float')
            pr_spend = pr_spend.groupby('Date',as_index=False).agg({'Amount':"sum"})
            pr_spend['Date'] = pd.to_datetime(pr_spend['Date'])
            pr_spend = pr_spend.assign(month=pr_spend['Date'].dt.month,
                                    year=pr_spend['Date'].dt.year)
            # Merge spend and grouped imps
            spendetails = pd.merge(pr_imps,pr_spend,how='left',on=['month','year'])
            spendetails = spendetails[['month','year','SEO Impact','Amount']]
            # Merge main dataframe to spendetails
            df = pd.merge(df,spendetails,how='left',on=['month','year'])
            df = df.rename(columns={'SEO Impact_x':'SEO Impact',
                                                    'SEO Impact_y':'grouped_impressions'})
            df['SEO Impact'] = df['SEO Impact'].astype('float64')
            df['grouped_impressions'] = df['grouped_impressions'].astype('float64')
            df = df.fillna(0)
            df = df.assign(SPEND=(df['SEO Impact'].astype('float')/df['grouped_impressions'].astype('float'))*df['Amount'].astype('float')) 
            #pr_original = pr_original.assign(SPEND="")
            #Try to remove the index
            pr = df.reset_index(drop=True)
            #Deduplicate dataset, optional flag: ignore_index=True
            pr = pr.drop_duplicates()
            #Fill NaNs with zeroes
            pr = pr.fillna(0)
            #Rename important columns so they concatenate properly later 
            pr = pr.rename(columns= {'Date':'DATE',
                                    'Title':'EVENT',
                                    'Readership':'IMPRESSIONS',
                                    'Shares':'CLICKS'})
            pr = pr.assign(FORMAT='PR',
                        PRODUCT='C1B')
            # PR
            pr = pr[['DATE'
                    ,'FORMAT'
                    ,'PRODUCT'
                    ,'EVENT'
                    ,'IMPRESSIONS'
                    ,'CLICKS'
                    ,'SPEND'
                    ,'Media Type'
                    ,'Media Outlet'
                    ,'Link'
                    ,'Author'
                    ,'Sentiment'
                    ,'Article Impact'
                    ,'SEO Impact'
                    ,'Tags'
                    ,'Country'
                    ,'State'
                    ,'City'
                    ,'Earned OR Syndicated'
                    ,'quality_site']]
            return(pr)
        
        def radio(self,source: str):
            target = assemblers().data_assembler(source,"(.*COB Data Dump.*\.xlsx)")
            raw = assemblers().data_maker(data=source,data_assembler=target,sheet='Radio')
            dedupe = raw.drop_duplicates()
            formatter = dedupe.assign(Format='Radio')
            rad = formatter.rename(columns={'Net Spend':'SPEND'
                                            ,'Purchased Impressions ':'IMPRESSIONS'
                                            ,'Format':'FORMAT'
                                            ,'MARKET':'Market'})
            rad = rad.assign(CLICKS=0,
                                EVENT='N/A',
                                PRODUCT='N/A')
            rad = rad[[  'DATE'
                        ,'FORMAT'
                        ,'PRODUCT'
                        ,'EVENT'
                        ,'IMPRESSIONS'
                        ,'CLICKS'
                        ,'SPEND'
                        ,'STA'
                        ,'AIR PROGRAM'
                        ,'DEMO #1'
                        ,'ESTIMATE NAME'
                        ,'LEN'
                        ,'Market'
                        ,'DP'
                        ,'Purchased'
                        ,'Purchased Ratings']]
            return rad
        
        def raiders(self,source: str):
            raiders = pd.read_excel(source+r"\Raiders YTD 2022.XLSX")
            ra = raiders[['Date','Sponsorship Impression','Event','Fixture','Match','Tool Location']]
            ra = ra.rename(columns={'Event':'Events',
                                    'Match':'Event',
                                    'Date':'DATE',
                                    'Tool Location':'PLACEMENT'})
            ra = ra.assign(Product = "RAIDERS",
                        Format = 'Sponsorship',
                        IMPRESSIONS= ra['Sponsorship Impression'])
            return ra

        def vgk(self,source: str):
            vg0 = pd.read_excel(source+r"\VGK YTD 21.XLSX",sheet_name='VGK Game Impressions')
            vg1 = pd.read_excel(source+r"\VGK YTD 22.XLSX",sheet_name='VGK Game Impressions')
            vg2 = pd.read_excel(source+r"\VGK YTD 23.XLSX",sheet_name='VGK Game Impressions')
            vgk = pd.concat([vg0,vg1,vg2])
            vgk = vgk.rename(columns={'Brands':'Events',
                                    'location':'PLACEMENT'})
            vg = vgk[['Events','Impressions (000)','PLACEMENT','DATE']]
            vg = vg.assign(Product="VGK")
            vg = vg.assign(Format='Sponsorship')
            vg['DATE'] = pd.to_datetime(vg['DATE'])
            vg = vg.assign(Event=vg['Events'].str.extract(r'\d+\-\d+\s(.*)'))
            vg = vg.assign(IMPRESSIONS=vg['Impressions (000)']*1000)
            return vg
        
        def wso(self,source: str):
            wso = pd.read_excel(source+r"\WSO.xlsx")
            ws = wso[['DATE','FORMAT','Fixture','location','Sum of Impressions','PRODUCT']]
            ws = ws.rename(columns={'Fixture':'Event',
                                    'location':'PLACEMENT',
                                # 'Sum of Impressions':'IMPRESSIONS',
                                    'FORMAT':'Format',
                                    'PRODUCT':'Product'})
            ws = ws.assign(IMPRESSIONS=ws['Sum of Impressions']*1000)
            return ws
        
        def wwe(self,source: str):
            wwe = pd.read_excel(source+r"\WWE YTD 22.xlsx")
            ww = wwe[['Format','Placement','Creative','Start Date','Delivered Impressions']]
            ww = ww.assign(Event=ww['Format']+ww['Placement']+" "+ww['Creative'])
            ww = ww.rename(columns={'Placement':'PLACEMENT',
                                    'Format':'tvformat'})
            ww = ww.assign(Product='WWE',
                        Event= ww['PLACEMENT']+" "+ww['Creative'],
                        IMPRESSIONS = ww['Delivered Impressions'],
                        Format='Sponsorship')
            return ww
    class dla:

        def __init__(self):
            pass

        def connect(self):
            # Thank you, Josh Perkins!
            # Where the file is being saved
            path = data.assumptions().data
            # Initializes an Outlook session and provides access to email-related functionalities through the MAPI namespace.
            outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI") 
            # Find the root folder by its name
            root_folder = None
            for folder in outlook.Folders:
                if folder.Name == "Brand Analytics Data Dump":
                    root_folder = folder
                    break
            if root_folder is None:
                print(f"{time.time()} >>> Root folder not found",file=open(log,'a'))
                exit()
            # Access the secondary inbox from the root folder
            inbox = root_folder.Folders.Item('Inbox')
            # Create a link to the folder where the emails will be moved
            downloaded = root_folder.Folders.Item('Downloaded')
            # Retrieves a collection of all email items within the specified folder
            messages = inbox.Items

            # Sender Keyword
            def save_attachments(keyword):
                # Create a list of messages to move
                messages_to_move = [message for message in messages if keyword.lower() in message.SenderEmailAddress.lower()]
                for message in messages_to_move:
                    attachments = message.Attachments
                    for attachment in attachments: 
                        if attachment.FileName.lower().endswith(('.xls', '.xlsx', '.csv')):
                            print(f"Saving attachment: {attachment.FileName}",file=open(log,'a'))
                            attachment.SaveAsFile(os.path.join(path, attachment.FileName))
                    # Changes the email to Read
                    if message.Unread:
                        message.Unread = False
                    # Move the message to the 'Downloaded' folder
                    message.Move(downloaded)
                    # print(f"Moving Email: {message.Subject}")

            keyword_to_search = "deutsch"
            save_attachments(keyword_to_search)

        def youtube(self):
            d = data.assumptions().data
            targets = assemblers().data_assembler(source=d,criteria='YouTube_C1B_DataDump.*')
            youtube_set = assemblers().data_maker(data=d,data_assembler=targets,skiprows=2)
            youtube_set = youtube_set.assign(FORMAT='Youtube')
            youtube_set = youtube_set.rename(columns={"Day":"DATE",
                                                    "Impr.":"IMPRESSIONS",
                                                    "Clicks":"CLICKS",
                                                    "Cost": "SPEND"})
            # extraction = youtube_set["Ad name"].str.extractall("([^_]*)").xs(4, level='match')
            # extraction = extraction.rename(columns = {extraction.columns[0]:"PRODUCT"})
            # youtube_set = youtube_set.join(extraction)
            youtube_set['DATE'] = pd.to_datetime(youtube_set['DATE'])
            return (youtube_set)
        
        def television(self):
            d = data.assumptions().data
            targets = assemblers().data_assembler(source=d,criteria='Deutsch-IDM Credit One Nat Cable.*')
            tv_set = assemblers().data_maker(data=d,data_assembler=targets)
            tv_set = tv_set.rename(columns={"Product":"PRODUCT",
                                    "Air Date":"DATE",
                                    "ISCI Length":"UnitLength",
                                    "Aired Prg Name": "EVENT",
                                    "Cmml Title":"Creative",
                                    "Demographic":"Demo",
                                    "Est AD18+ Full Impression":"IMPRESSIONS",
                                    "Spend":"SPEND",
                                    "Daypart":"DayPart"})
            tv_set = tv_set.assign(FORMAT = "TV")
            return tv_set

        def digital(self):
            d = data.assumptions().data
            targets = assemblers().data_assembler(source=d,criteria='DCM_C1B_DataDump.*')
            digital_set = assemblers().data_maker(data=d,data_assembler=targets)
            digital_set[['Start-End'
                    ,'Buy Type'
                    ,'Partner'
                    ,'Placement'
                    ,'Platform'
                    ,'Objective'
                    ,'WidthxHeight'
                    ,'Buying Tool'
                    ,'Geography'
                    ,'Purchase Type'
                    ,'Targeting'
                    ,'Audience'
                    ,'Marketplace'
                    ,'Pricing'
                    ,'test control']] = digital_set['Placement'].str.split('_', expand=True)
            digital_set = digital_set.rename(columns={"Day of Date":"DATE",
                                                "Impressions":"IMPRESSIONS",
                                                "Spend":"SPEND",
                                                "Creative":"EVENT"})
            digital_set = digital_set.assign(FORMAT = np.where(digital_set["Placement"] == "Video", "CTV", "Display"))
            # modified
            return digital_set
        
        def adelaide(self):
            data.dla().connect()
            targets = assemblers().data_assembler(source=data.assumptions().data,criteria='Adelaide.*')
            adelaide_DRTV = assemblers().data_maker(data=data.assumptions().data,data_assembler=targets,sheet="DRTV")
            adelaide_CTV_Display_Social = assemblers().data_maker(data=data.assumptions().data,data_assembler=targets,sheet="CTV_Display_Social")
            adelaide_DRTV = adelaide_DRTV.assign(FORMAT="TV")
            adelaide_CTV_Display_Social["FORMAT"] = np.where(adelaide_CTV_Display_Social["Channel"].isin(["Facebook","Instagram"]),"Social",adelaide_CTV_Display_Social["Channel"])
            adelaide_DRTV['ISCI Length'] = adelaide_DRTV['ISCI Length'].astype('str')
            adelaide_DRTV = adelaide_DRTV.rename(columns={"Air Date":"DATE","Daypart ":"DAYPART","Average AU":"AVG_AU","Impressions":"IMPRESSIONS","Total AU":"TOTAL_AU","ISCI Length":"DIMENSIONS"})
            adelaide_CTV_Display_Social = adelaide_CTV_Display_Social.rename(columns={"Date Date":"DATE","Daypart (Local Time)":"DAYPART","Average AU":"AVG_AU","AU Measurable Impressions":"IMPRESSIONS","Total AU":"TOTAL_AU","Ad Format (Final)":"DIMENSIONS"})
            adelaide_data = pd.concat([adelaide_DRTV,adelaide_CTV_Display_Social])
            adelaide_data.to_csv(data.assumptions().product+r"\adelaide_data.csv")
            return (data.assumptions().product, adelaide_DRTV, adelaide_CTV_Display_Social,) # <<< Thats's a friggin tuple

        
        def VideoAmpDataProcessor(self):
            d = data.assumptions()

            targets = assemblers().data_assembler(source=data.assumptions().data,criteria='.*VideoAmp.*')

            Monthly_by_Partner = assemblers().data_maker(data=data.assumptions().data,data_assembler=targets,sheet="Monthly by Partner")

            site_visits_df = assemblers().data_maker(data=data.assumptions().data,data_assembler=targets,sheet="Site Visits")
            # Remove blank leading column
            site_visits_clean = site_visits_df.dropna(axis=1, how='all')
            # Delete rows until "Broadcast Month" is reached
            for num, row in enumerate(site_visits_clean.values):
                if row[0] == 'Broadcast Month':
                    break
                else:
                    site_visits_clean = site_visits_clean.drop(num)
            # When you remove the first couple of rows it means the first item in the index is 2, not zero. Hence the reset_index.
            site_visits_clean = site_visits_clean.reset_index(drop=True)
            # Make the first row the headers
            new_headers = site_visits_clean.iloc[0]
            site_visits_clean = site_visits_clean[1:]
            site_visits_clean.columns = new_headers
            Site_Visits = site_visits_clean 

            monthly_site_visits_by_partner = pd.merge(Site_Visits,Monthly_by_Partner,how="left",on = ["Broadcast Month","Impressions","Channel","Station","Network"])
            monthly_site_visits_by_partner['Broadcast Month'] = monthly_site_visits_by_partner['Broadcast Month'].str.extract("M\s(\d{2}.*)")
            monthly_site_visits_by_partner['Broadcast Month'] = pd.to_datetime(monthly_site_visits_by_partner['Broadcast Month'])

            monthly_site_visits_by_partner.to_csv(d.product+r"\monthly_site_visits_by_partner.csv")

            Meta_Data_by_Placement = assemblers().data_maker(data=data.assumptions().data,data_assembler=targets,sheet="Meta Data by Placement")
            Meta_Data_by_Placement.to_csv(d.product+r"\Meta_Data_by_Placement"+".csv")

            return (monthly_site_visits_by_partner, Meta_Data_by_Placement)

        
    class social:

        def __init__(self):
            pass

        def assumptions(self):
            namekey = os.environ['USERNAME']
            self.file_string = "C:\\Users\\{namekey}\\AppData\\Roaming\\Python\\Python310\\site-packages\\facebook_business\\config.json".format(namekey=namekey)

            try:
                self.config = open(self.file_string)
                self.d = json.load(self.config)
                print(f"{time.time()} >>> Configuration successfully loaded...",flush=True)
            except:
                print(f"{time.time()} >>> Make sure you've installed the SDK and that it is in the right file path.",flush=True)
            self.app_id = self.d['app_id']
            self.app_secret = self.d['app_secret']
            self.accounts = self.d['accounts']
            self.access_token = self.d['access_token']
            self.initial_access_token = self.d['initial_access_token']

            self.file_destination = r"N:\Marketing Analytics\Brand Analysis\04 - Dashboards\Brand Spend Data Model\Source Data"
            self.file_name = r"\C1B Social Data - {}.xlsx".format(datetime.today().date())

            self.old_file_list = os.listdir(self.file_destination)
            self.old_social_file_list = []
            for i in self.old_file_list:
                if i.__contains__("C1B Social Data -"):
                    self.old_social_file_list.append(i)
            if len(self.old_social_file_list) ==0:
                self.start_date = '2024-01-01'
            else:
                self.previously_retrieved_data = pd.DataFrame()
                for i in self.old_social_file_list:
                    x = pd.read_excel(self.file_destination+"\\"+i)
                    self.previously_retrieved_data = pd.concat([self.previously_retrieved_data,x])
                    os.remove(self.file_destination+"\\"+i)
                self.start_date = self.previously_retrieved_data["Day"]
            self.end_date = datetime.today() - timedelta(days=1)
            self.datelist = pd.date_range(self.start_date,self.end_date)

        class authorizer:
            def __init__(self):
                self.d = data.social()
                self.d.assumptions()
                self.access_token = self.d.access_token
                self.app_id = self.d.app_id
                self.app_secret = self.d.app_secret
                self.file_string = self.d.file_string
                self.initial_access_token = self.d.initial_access_token
                self.access_token_sorter()
                
        #
            def get_fb_token(self, access_token: str = None) -> requests.models.Response:
                url = 'https://graph.facebook.com/v18.0/oauth/access_token'
                payload = {
                    'grant_type': 'fb_exchange_token',
                    'client_id': self.app_id,
                    'client_secret': self.app_secret,
                    'fb_exchange_token': access_token
                }
                response = requests.get(url, params=payload)
                return response
        #
            def access_checker(self, access_token: str = None) -> bool:
                connection = FacebookAdsApi.init(access_token=access_token,app_secret=self.app_secret)
                try:
                    calling = connection.call('GET',"https://graph.facebook.com/v18.0/me?fields=id%2Cname&access_token={}".format(access_token))._http_status
                except:
                    calling = 400
                if calling == 200:
                    result = True
                else:
                    result = False
                return result
        #
            def token_writer(self, access_token: str) -> None:
                config = open(self.file_string)
                d = json.load(config)
                d['access_token'] = access_token
                with open(self.file_string,"w") as outfile:
                    json.dump(d, outfile)
        #
            def access_token_sorter(self) -> str:
                if len(self.access_token) > 1:
                    if self.access_checker(access_token=self.access_token):
                        print(f"{time.time()} >>> Option 1: Long-term access token available and valid.",file=open(log,'a'))
                        access_token = self.access_token
                        days = 60
                    else:
                        print(f"{time.time()} >>> Option 2: Long-term access token available, but invalid.",file=open(log,'a'))
                        new_access_token = input("Please enter a new access token.")
                        response = self.get_fb_token(access_token=new_access_token)
                        access_token = response.json()['access_token']
                        self.token_writer(access_token=access_token)
                        days = response.json()['expires_in'] / 60 / 60 / 24
                else:
                    if self.access_checker(access_token=self.initial_access_token):
                        print(f"{time.time()} >>> Option 3: No long-term access token, initial access token used to create long-term access token.",file=open(log,'a'))
                        response = self.get_fb_token(access_token=self.initial_access_token)
                        access_token = response.json()['access_token']
                        self.token_writer(access_token=access_token)
                        days = response.json()['expires_in'] / 60 / 60 / 24
                    else:
                        print(f"{time.time()} >>> option 4: No valid access tokens availble, please enter new token.",file=open(log,'a'))
                        new_access_token = input("Please enter a new access token.")
                        response = self.get_fb_token(access_token=new_access_token)
                        access_token = response.json()['access_token']
                        self.token_writer(access_token=access_token)
                        days = response.json()['expires_in'] / 60 / 60 / 24
                
                print(f"This access_token expires in {str(days)} days.",file=open(log,'a'))
                return access_token

        class report:
            def __init__(self):
                pass
            def builder(self, data = None, target_date: datetime = None):
            # Initial access to API
                start_date = str(target_date)[0:10]
                end_date = target_date + timedelta(days=1)
                end_date = str(end_date)[0:10]
                print(start_date,type(start_date),end_date)
                output = pd.DataFrame()
                for i in data.accounts:
                    camp = AdAccount(i).get_campaigns(params={'time_range':{'since':start_date,'until':end_date}})
                    l = []
                    while len(camp._queue) > 1 and camp._total_count > 1:
                        for i in camp._queue:
                            l.append(i['id'])
                        camp.load_next_page()
                    # l now has a list of campaigns to iterate through
                    # Get the insights out of a campaign
                    c_fields = [
                                'campaign_name',
                                'adset_name',
                                'account_name',
                                #'effective_status', Not allowed in insights for some reason
                                'ad_name',
                                'reach',
                                'spend',
                                'impressions',
                                'clicks',
                                'cpm',
                                'cpc',
                                'created_time',
                                'updated_time',
                                'date_start',
                                'date_stop',
                                'frequency', 
                                'dda_results', 
                                'place_page_name', 
                                'total_postbacks_detailed_v4'
                                ]
                    c_params = {
                                'level' : 'ad',
                                'breakdowns' :  [        
                                    'device_platform',
                                    'platform_position',
                                    'publisher_platform'
                                    ],
                                'time_range':{'since':start_date,'until':end_date},
                                'time_increment': '1'
                                }
                    # Loop through campaigns and compile dataframes into output variable
                    for c in l:
                        csight = Campaign(c).get_insights(fields = c_fields, params = c_params)
                        # print("Working on campaign "+str(c),flush=True)
                        while len(csight._queue) > 1:
                            print(f"{time.time()} >>> Length of queue: "+str(len(csight._queue)))
                            for i in csight._queue:
                                x = dict(i)
                                x = pd.DataFrame(x,index =[0])
                                output = pd.concat([output,x])
                                time.sleep(3)
                            csight.load_next_page()
                return output

        def run(self):
            auth = data.social.authorizer()
            for i in auth.d.datelist:
                print(i,type(i))
                load = data.social.report().builder(data=auth.d,target_date=i)
                load.to_excel(auth.d.file_destination+auth.d.file_name)
            return load
        
    def rudimentary_social(self):
            t = assemblers().data_assembler(source=data.assumptions().data,criteria="Meta.*")
            tt = assemblers().data_maker(data=data.assumptions().data,data_assembler=t)
            tt = tt.rename(columns={"Impressions":"IMPRESSIONS",
                                    "Day":"DATE",
                                    "Amount spent (USD)":"SPEND"})
            tt['DATE'] = pd.to_datetime(tt['DATE'],format='ISO8601')
            tt = tt.assign(FORMAT="Social")

            return tt

    # class application_survey:
    #     def __init__(self):
    #         pass
    #     def ingest(self) -> parquet:
    #         app_survey_data = pd.read_parquet(r"N:\Marketing Analytics\Brand Analysis\04 - Dashboards\Application Survey\01 - Data\01 - Dashboard")

