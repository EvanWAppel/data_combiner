import pandas as pd
import utils
import numpy as np
import openpyxl as xl
import re
import os.path

def tv(source: str):
    tv_set = utils.data_framer(source = source
                               , files = utils.file_gatherer(source=source,criteria='Deutsch-IDM Credit One Nat Cable.*')
                               , skiprows=None
                               , sheets = None)
    tv_set = tv_set.rename(columns={"Product":"PRODUCT",
                            "Air Date":"DATE",
                            "ISCI Length":"UnitLength",
                            "Aired Prg Name": "EVENT",
                            "Cmml Title":"Creative",
                            "Demographic":"Demo",
                            "Est AD18+ Full Impression":"IMPRESSIONS",
                            "Spend":"SPEND",
                            "Daypart":"DayPart"})
    tv_set = tv_set.assign(FORMAT = "TV")
    utils.entry(f"{len(tv_set)} rows loaded from TV.")
    return tv_set

def digital(source: str):
    digital_set = utils.data_framer(source = source
                               , files = utils.file_gatherer(source=source,criteria='DCM_C1B_DataDump.*')
                               , skiprows=None
                               , sheets = None)
    digital_set[['Start-End'
            ,'Buy Type'
            ,'Partner'
            ,'Placement'
            ,'Platform'
            ,'Objective'
            ,'WidthxHeight'
            ,'Buying Tool'
            ,'Geography'
            ,'Purchase Type'
            ,'Targeting'
            ,'Audience'
            ,'Marketplace'
            ,'Pricing'
            ,'test control']] = digital_set['Placement'].str.split('_', expand=True)
    digital_set = digital_set.rename(columns={"Day of Date":"DATE",
                                        "Impressions":"IMPRESSIONS",
                                        "Spend":"SPEND",
                                        "Creative":"EVENT"})
    digital_set = digital_set.assign(FORMAT = np.where(digital_set["Placement"] == "Video", "CTV", "Display"))
    # modified
    utils.entry(f"{len(digital_set)} rows loaded from DIGITAL.")
    return digital_set

def pr(source:str):
    files = utils.file_gatherer(source=source,criteria=r".*Credit One Bank Coverage Report.*")
    targets = []
    for i in files:
        for j in xl.load_workbook(os.path.join(source,i)).sheetnames:
            if re.search(".*\d{4}",j):
                targets.append((i,j))
    df = pd.DataFrame()
    for k in targets:
        sdf = utils.data_framer(source=source
                                ,files=[k[0]]
                                ,sheets=[k[1]]
                                ,skiprows=2)
        df = pd.concat([df,sdf])
    # df['Date'] = pd.to_datetime(df['Date'])
    # # Create new columns for month and year
    # df = df.assign(month=df['Date'].dt.month,
    #                year=df['Date'].dt.year)
   
    # PR
    pr = df.rename(columns={"Date":"DATE",
                            "Outlet":"PLACEMENT",
                            "Headline":"EVENT",
                            "UVM":"IMPRESSIONS"})
    
    pr = pr.assign(FORMAT="PR")

    utils.entry(f"{len(pr)} rows loaded from MLB.")
    return(pr)

def social_no_api(source: str):
    soc = utils.data_framer(source=source
                            , files=utils.file_gatherer(source=source,criteria="social-ingestion-protocol.*"))
    soc = soc.rename(columns={"Impressions":"IMPRESSIONS",
                            "Day":"DATE",
                            "Amount spent (USD)":"SPEND",
                            "Clicks (all)":"CLICKS"})
    soc['DATE'] = pd.to_datetime(soc['DATE'],format='ISO8601')
    soc = soc.assign(FORMAT="Social"
                     ,Video_plays="-")
    soc['ACCOUNT'] = ["DLA" if x == "Facebook Credit One Ad buys" else "C1B" for x in soc['Account name']]

    # soc =soc[['DATE'
    #         ,'FORMAT'
    #         #,'PRODUCT'
    #         #,'EVENT'
    #         #,'PLACEMENT'
    #         ,'IMPRESSIONS'
    #         ,'CLICKS'
    #         ,'SPEND'
    #         ,'ACCOUNT']]
    utils.entry(f"{len(soc)} rows loaded from soc.")
    return soc

def coco(source: str):
    coco = utils.data_framer(source = source
                        , files = utils.file_gatherer(source = source,
                                                    criteria = 'COCO.*')
                        , sheets = ["COCO FINAL"]
                        , skiprows = None)
    co = coco[['DATE','Format','Events','Locations','IMPRESSIONS']]
    co = co.rename(columns={'Events':'Event',
                            'Locations':'PLACEMENT'})
    co = co.assign(Product="COCO")
    utils.entry(f"{len(co)} rows loaded from COCO.")
    return co

def mlb(source: pd.DataFrame):
    mlba = utils.data_framer(source = source
                        , files = utils.file_gatherer(source = source,
                                                    criteria = 'mlb.*')
                        , sheets = ["Sheet1"]
                        , skiprows = None)
    spend2023 = 864690
    mlba = mlba.assign(IMPRESSIONS=mlba['Sponsorship Impressions (000)']*1000,
                    FORMAT = 'TV',
                    PRODUCT = 'MLB')
    mlba = mlba.assign(tot_imp = np.sum(mlba['IMPRESSIONS']))
    mlba = mlba.assign(SPEND = (spend2023/mlba['tot_imp'])*mlba['IMPRESSIONS'])
    utils.entry(f"{len(mlba)} rows loaded from MLB.")
    return mlba

def nascar(source: str):
    nascar = utils.data_framer(source = source
                        , files = utils.file_gatherer(source = source,
                                                    criteria = '.*Nascar.*')
                        , sheets = ["Nascar Race Level"]
                        , skiprows = None)
    na = nascar[['Events','Exposure Impressions (000)','DATE']]
    na = na.assign(Product="NASCAR")
    na = na.assign(Format='Sponsorship')
    na['DATE'] = pd.to_datetime(na['DATE'])
    na = na.assign(Event=na['Events'].str.extract(r'\d+\-\d+\s(.*)'))
    na = na.assign(IMPRESSIONS=na['Exposure Impressions (000)']*1000)
    utils.entry(f"{len(na)} rows loaded from NASCAR.")
    return na

def vgk(source: str):
    vgk = utils.data_framer(source=source
                            ,files=utils.file_gatherer(source=source
                                                        ,criteria=r'.*VGK YTD.*')
                            ,sheets = ['VGK Game Impressions']
                            ,skiprows=None)
    vgk = vgk.rename(columns={'Brands':'Events',
                            'location':'PLACEMENT'})
    vg = vgk[['Events','Impressions (000)','PLACEMENT','DATE']]
    vg = vg.assign(Product="VGK")
    vg = vg.assign(Format='Sponsorship')
    vg['DATE'] = pd.to_datetime(vg['DATE'])
    vg = vg.assign(Event=vg['Events'].str.extract(r'\d+\-\d+\s(.*)'))
    vg = vg.assign(IMPRESSIONS=vg['Impressions (000)']*1000)
    return vg

def raiders(source: str):
        raiders = utils.data_framer(source=source
                                    ,files=utils.file_gatherer(source=source,criteria=".*Raider.*")
                                    ,sheets=["Data"]
                                    ,skiprows=None)
        raiders = pd.read_excel(source+r"\Raiders YTD 2022.XLSX")
        ra = raiders[['Date','Sponsorship Impression','Event','Fixture','Match','Tool Location']]
        ra = ra.rename(columns={'Event':'Events',
                                'Match':'Event',
                                'Date':'DATE',
                                'Tool Location':'PLACEMENT'})
        ra = ra.assign(Product = "RAIDERS",
                    Format = 'Sponsorship',
                    IMPRESSIONS= ra['Sponsorship Impression'])
        utils.entry(f"{len(ra)} rows loaded from Raiders.")
        return ra

def wso(source: str):
    wso = pd.read_excel(source+r"\WSO.xlsx")
    ws = wso[['DATE','FORMAT','Fixture','location','Sum of Impressions','PRODUCT']]
    ws = ws.rename(columns={'Fixture':'Event',
                            'location':'PLACEMENT',
                        # 'Sum of Impressions':'IMPRESSIONS',
                            'FORMAT':'Format',
                            'PRODUCT':'Product'})
    ws = ws.assign(IMPRESSIONS=ws['Sum of Impressions']*1000)
    return ws

def wwe(source: str):
    wwe = utils.data_framer(source = source
                           ,files = utils.file_gatherer(source=source,criteria=".*WWE YTD.*")
                           ,sheets=["WWE PBI"]
                           ,skiprows=None)
    ww = wwe[['Format','Placement','Creative','Start Date','Delivered Impressions']]
    ww = ww.assign(Event=ww['Format']+ww['Placement']+" "+ww['Creative'])
    ww = ww.rename(columns={'Placement':'PLACEMENT',
                            'Format':'tvformat'})
    ww = ww.assign(Product='WWE',
                Event= ww['PLACEMENT']+" "+ww['Creative'],
                IMPRESSIONS = ww['Delivered Impressions'],
                Format='Sponsorship')
    return ww

def youtube(source:str):
    youtube_set = utils.data_framer(source=source
                                    , files=utils.file_gatherer(source=source,criteria='YouTube_C1B_DataDump.*')
                                    , sheets = None
                                    , skiprows = 2)
    youtube_set = youtube_set.assign(FORMAT='Youtube')
    youtube_set = youtube_set.rename(columns={"Day":"DATE",
                                            "Impr.":"IMPRESSIONS",
                                            "Clicks":"CLICKS",
                                            "Cost": "SPEND"})
    youtube_set['DATE'] = pd.to_datetime(youtube_set['DATE'])
    return (youtube_set)

def adelaide(source: str,target: str):
    adelaide_DRTV = utils.data_framer(source=source
                                    , files=utils.file_gatherer(source=source,criteria='Adelaide.*')
                                    , sheets = ["DRTV"]
                                    , skiprows = None)
    adelaide_CTV_Display_Social = utils.data_framer(source=source
                                    , files=utils.file_gatherer(source=source,criteria='Adelaide.*')
                                    , sheets = ["CTV_Display_Social"]
                                    , skiprows = None)
    adelaide_DRTV = adelaide_DRTV.assign(FORMAT="TV")
    adelaide_CTV_Display_Social["FORMAT"] = np.where(adelaide_CTV_Display_Social["Channel"].isin(["Facebook","Instagram"]),"Social",adelaide_CTV_Display_Social["Channel"])
    adelaide_DRTV['ISCI Length'] = adelaide_DRTV['ISCI Length'].astype('str')
    adelaide_DRTV = adelaide_DRTV.rename(columns={"Air Date":"DATE","Daypart ":"DAYPART","Average AU":"AVG_AU","Impressions":"IMPRESSIONS","Total AU":"TOTAL_AU","ISCI Length":"DIMENSIONS"})
    adelaide_CTV_Display_Social = adelaide_CTV_Display_Social.rename(columns={"Date Date":"DATE","Daypart (Local Time)":"DAYPART","Average AU":"AVG_AU","AU Measurable Impressions":"IMPRESSIONS","Total AU":"TOTAL_AU","Ad Format (Final)":"DIMENSIONS"})
    adelaide_data = pd.concat([adelaide_DRTV,adelaide_CTV_Display_Social])
    adelaide_data.to_csv(os.path.join(target+r"adelaide_data.csv"))
    return (target, adelaide_DRTV, adelaide_CTV_Display_Social,) # <<< Thats's a friggin tuple

def videoamp(source: str,target: str):
    # files = utils.file_gatherer(source=source,criteria='YouTube_C1B_DataDump.*')
    # Monthly_by_Partner = utils.data_framer(source=source
    #                                 , files=files
    #                                 , sheets = ["Monthly by Partner"]
    #                                 , skiprows = None)
    # site_visits_df = utils.data_framer(source=source
    #                                 , files=files
    #                                 , sheets = ["Site Visits"]
    #                                 , skiprows = None)
    # Meta_Data_by_Placement = utils.data_framer(source=source
    #                                 , files=files
    #                                 , sheets = ["Meta Data by Placement"]
    #                                 , skiprows = None)
    # # Remove blank leading column
    # site_visits_clean = site_visits_df.dropna(axis=1, how='all')
    # # Delete rows until "Broadcast Month" is reached
    # for num, row in enumerate(site_visits_clean.values):
    #     if row[0] == 'Broadcast Month':
    #         break
    #     else:
    #         site_visits_clean = site_visits_clean.drop(num)
    # # When you remove the first couple of rows it means the first item in the index is 2, not zero. Hence the reset_index.
    # site_visits_clean = site_visits_clean.reset_index(drop=True)
    # # Make the first row the headers
    # new_headers = site_visits_clean.iloc[0]
    # site_visits_clean = site_visits_clean[1:]
    # site_visits_clean.columns = new_headers
    # Site_Visits = site_visits_clean 
    # monthly_site_visits_by_partner = pd.merge(Site_Visits,Monthly_by_Partner,how="left",on = ["Broadcast Month","Impressions","Channel","Station","Network"])
    # monthly_site_visits_by_partner['Broadcast Month'] = monthly_site_visits_by_partner['Broadcast Month'].str.extract("M\s(\d{2}.*)")
    # monthly_site_visits_by_partner['Broadcast Month'] = pd.to_datetime(monthly_site_visits_by_partner['Broadcast Month'])
    # monthly_site_visits_by_partner.to_csv(os.path.join(target,r"monthly_site_visits_by_partner.csv"))
    # Meta_Data_by_Placement.to_csv(os.path.join(target,r"Meta_Data_by_Placement",".csv"))
    monthly_site_visits_by_partner = pd.DataFrame()
    Meta_Data_by_Placement = pd.DataFrame()
    return (monthly_site_visits_by_partner, Meta_Data_by_Placement)